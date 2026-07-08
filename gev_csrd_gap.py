"""
GE Vernova — CSRD / ESRS Gap Analysis & Double Materiality Assessment

Data source : GE Vernova Sustainability Report 2025
Standard    : ESRS Set 1 (2023), CSRD (EU) 2022/2464, ESRS 1 Art. 3 (DMA)
Question    : If GE Vernova had to file under CSRD/ESRS — where are the gaps?
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.gridspec as gridspec
from matplotlib.patches import FancyBboxPatch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

DARK_TEXT = "#1A1A1A"
MID_TEXT  = "#666666"
GRID_COL  = "#E8E8E8"
GEV_BLUE  = "#005EB8"
GEV_GREEN = "#00A651"
GEV_RED   = "#E53935"
GEV_ORANGE= "#FF6F00"

plt.rcParams.update({
    "font.family":       "DejaVu Sans",
    "font.size":         10,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "axes.edgecolor":    "#444444",
    "axes.linewidth":    0.8,
})


# ════════════════════════════════════════════════════════════════════════════
# DATA — ESRS gap assessment based on GE Vernova 2025 report analysis
# ════════════════════════════════════════════════════════════════════════════

ESRS = pd.DataFrame({
    "topic": [
        "E1 — Climate Change",
        "E2 — Pollution",
        "E3 — Water & Marine Resources",
        "E4 — Biodiversity & Ecosystems",
        "E5 — Resource Use & Circular Economy",
        "S1 — Own Workforce",
        "S2 — Value Chain Workers",
        "S3 — Affected Communities",
        "S4 — Consumers & End-Users",
        "G1 — Business Conduct",
    ],
    "pillar": ["E","E","E","E","E","S","S","S","S","G"],
    "current_score": [7.5, 3.0, 4.0, 2.5, 6.0, 7.0, 4.5, 3.5, 1.0, 5.5],
    "gap_score":     [2.5, 7.0, 6.0, 7.5, 4.0, 3.0, 5.5, 6.5, 9.0, 4.5],
    "disclosed": [True, False, False, False, True, True, False, False, False, True],
    "color": ["#1565C0","#6A1B9A","#00838F","#2E7D32","#E65100",
              "#C62828","#AD1457","#F57F17","#558B2F","#37474F"],
    "what_gev_discloses": [
        "Scope 1+2+3, carbon intensity, avoided emissions, 2030 target, energy consumption",
        "Spills & releases count, air/wastewater exceedances, environmental penalties",
        "Water risk review completed 2025; water savings initiatives reported",
        "Biodiversity Principles launched 2024; no site-level data",
        "4R circularity framework, LCA/EPD coverage, product circularity %",
        "TRIR, DAWR, fatalities, engagement score, headcount by region",
        "SRG supply chain audits, human rights statement, supplier code of conduct",
        "GE Vernova Foundation; community programs; philanthropy spend",
        "Product safety philosophy; Life Saving Rules; quality management",
        "Code of conduct, anti-bribery policy, ethics hotline, audit findings",
    ],
    "key_gap": [
        "No double materiality assessment; no EU Taxonomy alignment; transition plan lacks financial effect disclosure",
        "No pollutant-specific KPIs (NOx, SOx, PM); no pollution prevention targets by substance",
        "No withdrawal/consumption by water stress area; no ESRS E3 standard KPIs disclosed",
        "No site-level biodiversity assessment; no measurable biodiversity targets set",
        "No absolute waste/resource consumption targets; no circular revenue disclosure",
        "No gender pay gap; no health surveillance; no work-life balance disclosures",
        "No supplier GHG footprint; EUDR not addressed; no tier 2+ supplier coverage",
        "No community impact measurement; no grievance mechanism statistics",
        "No product complaint data; no cybersecurity incident disclosure; no customer satisfaction KPIs",
        "No supply chain ethics KPIs; no lobbying spend disclosure",
    ],
    "priority": ["High","Medium","Medium","Low","Medium",
                 "High","Medium","Low","Low","Medium"],
})

# Double materiality scores
DMA = pd.DataFrame({
    "topic":     ["E1 Climate","E2 Pollution","E3 Water","E4 Biodiversity","E5 Circularity",
                  "S1 Workforce","S2 Value Chain","S3 Communities","S4 Consumers","G1 Conduct"],
    "impact":    [4.8, 3.5, 3.0, 2.8, 3.8, 4.2, 3.5, 3.0, 2.5, 3.2],
    "financial": [4.5, 2.8, 3.2, 2.5, 3.5, 3.8, 3.0, 2.5, 2.8, 4.0],
    "disclosed": [True,False,False,False,True,True,False,False,False,True],
    "color":     ["#1565C0","#6A1B9A","#00838F","#2E7D32","#E65100",
                  "#C62828","#AD1457","#F57F17","#558B2F","#37474F"],
})

print("Data loaded.")
print(f"ESRS topics assessed     : {len(ESRS)}")
print(f"Overall readiness score  : {ESRS['current_score'].mean():.1f}/10")
print(f"Topics with score < 4    : {(ESRS['current_score']<4).sum()}")
print(f"#1 missing requirement   : Double Materiality Assessment (ESRS 1, Art. 3)")


# ════════════════════════════════════════════════════════════════════════════
# FIGURE 1 — ESRS GAP ANALYSIS
# ════════════════════════════════════════════════════════════════════════════

fig1, (ax_l, ax_r) = plt.subplots(1, 2, figsize=(18, 9),
                                   gridspec_kw={"width_ratios":[1.2, 1]})
fig1.patch.set_facecolor("white")
fig1.suptitle(
    "Figure 1.  CSRD / ESRS Gap Analysis — GE Vernova (2025)\n"
    "If GE Vernova Had to File Under ESRS Set 1 — Where Are the Gaps?",
    fontsize=13, fontweight="bold", y=0.99, color=DARK_TEXT)

# Left: stacked bar current vs gap
ax_l.set_facecolor("#F8F8F8")
y = np.arange(len(ESRS))

bars_cur = ax_l.barh(y, ESRS["current_score"], height=0.55,
                     color=ESRS["color"], edgecolor="white",
                     linewidth=0.8, alpha=0.85, label="Currently disclosed")
bars_gap = ax_l.barh(y, ESRS["gap_score"], height=0.55,
                     left=ESRS["current_score"],
                     color=ESRS["color"], edgecolor="white",
                     linewidth=0.8, alpha=0.22, hatch="///",
                     label="Gap to ESRS compliance")

for bar, cur, gap in zip(bars_cur, ESRS["current_score"], ESRS["gap_score"]):
    ax_l.text(cur / 2, bar.get_y() + bar.get_height()/2,
              f"{cur:.1f}", va="center", ha="center",
              fontsize=8.5, fontweight="bold", color="white")
    ax_l.text(cur + gap + 0.1, bar.get_y() + bar.get_height()/2,
              f"Gap: {gap:.1f}", va="center", ha="left",
              fontsize=8, color=DARK_TEXT)

ax_l.axvline(10, color="#888", linewidth=1.2,
             linestyle="--", alpha=0.5)
ax_l.text(10.05, -0.7, "Full compliance (10)",
          fontsize=8, color="#555", style="italic")

ax_l.set_yticks(y)
ax_l.set_yticklabels(ESRS["topic"], fontsize=9, color=DARK_TEXT)
ax_l.set_xlabel("ESRS Disclosure Score (0–10)", fontsize=10, color=DARK_TEXT)
ax_l.set_xlim(0, 12.5)
ax_l.xaxis.grid(True, color=GRID_COL, linewidth=0.7, zorder=0)
ax_l.set_axisbelow(True)
ax_l.legend(frameon=False, fontsize=9, loc="lower right")
ax_l.set_title("(A)  Current ESRS Disclosure Score vs. Compliance Gap",
               fontsize=11, fontweight="bold", pad=10, loc="left", color=DARK_TEXT)

# Right: key gaps table
ax_r.set_facecolor("white")
ax_r.axis("off")
ax_r.set_title("(B)  Primary CSRD Gap per Topic",
               fontsize=11, fontweight="bold", pad=10, loc="left", color=DARK_TEXT)

col_x = [0.0, 0.25, 0.32, 0.97]
col_w = [0.24, 0.06, 0.65, 0]
hdrs  = ["Topic", "Score", "Primary Gap"]
row_h = 0.082
start_y = 0.90

for ch, cx, cw in zip(hdrs, col_x, col_w):
    rect = FancyBboxPatch((cx, start_y), cw, row_h,
                          boxstyle="round,pad=0.003",
                          facecolor="#1E3A5F", edgecolor="white",
                          linewidth=0.5, transform=ax_r.transAxes)
    ax_r.add_patch(rect)
    ax_r.text(cx + cw/2, start_y + row_h/2, ch,
              ha="center", va="center", fontsize=9,
              fontweight="bold", color="white",
              transform=ax_r.transAxes)

for i, (_, row) in enumerate(ESRS.iterrows()):
    y_row = start_y - (i+1) * row_h
    bg = "#F5F5F5" if i % 2 == 0 else "white"
    rect = FancyBboxPatch((col_x[0], y_row), 0.97, row_h-0.003,
                          boxstyle="round,pad=0.003",
                          facecolor=bg, edgecolor="#EEEEEE",
                          linewidth=0.3, transform=ax_r.transAxes)
    ax_r.add_patch(rect)

    ax_r.text(col_x[0]+0.01, y_row+row_h/2,
              row["topic"].split("—")[0].strip(),
              va="center", fontsize=8.5, fontweight="bold",
              color=row["color"], transform=ax_r.transAxes)

    sc = row["current_score"]
    sc_c = (GEV_GREEN if sc >= 7 else GEV_ORANGE if sc >= 4 else GEV_RED)
    ax_r.text(col_x[1]+col_w[1]/2, y_row+row_h/2,
              f"{sc:.1f}", ha="center", va="center",
              fontsize=9, fontweight="bold", color=sc_c,
              transform=ax_r.transAxes)

    gap_txt = row["key_gap"][:62]+"…" if len(row["key_gap"]) > 62 else row["key_gap"]
    ax_r.text(col_x[2]+0.01, y_row+row_h/2, gap_txt,
              va="center", fontsize=7.2, color=DARK_TEXT,
              transform=ax_r.transAxes)

overall = ESRS["current_score"].mean()
ax_r.text(0.5, 0.03,
    f"Overall ESRS readiness: {overall:.1f}/10  |  "
    "Double Materiality Assessment is the #1 missing CSRD requirement",
    ha="center", fontsize=8.5, color=GEV_RED,
    fontweight="bold", transform=ax_r.transAxes)

fig1.text(0.01, 0.01,
    "Scoring: 0–10 per topic based on sub-requirement completeness. "
    "ESRS Set 1 (2023) topical standards. "
    "DMA = Double Materiality Assessment required under ESRS 1, Art. 3.",
    fontsize=7.5, color=MID_TEXT, style="italic")

plt.tight_layout(rect=[0, 0.04, 1, 0.95])
plt.savefig(f"{OUTPUT_DIR}/fig1_esrs_gap.png",
            dpi=150, bbox_inches="tight", facecolor="white")
plt.show(); plt.close()
print("Figure 1 saved.")


# ════════════════════════════════════════════════════════════════════════════
# FIGURE 2 — DOUBLE MATERIALITY MATRIX
# ════════════════════════════════════════════════════════════════════════════

fig2, ax2 = plt.subplots(figsize=(12, 10))
fig2.patch.set_facecolor("white")
ax2.set_facecolor("white")

ax2.fill_between([1,3,3,1],[3,3,5.3,5.3], alpha=0.05, color="#FF7043")
ax2.fill_between([3,5.3,5.3,3],[1,1,3,3], alpha=0.05, color="#42A5F5")
ax2.fill_between([3,5.3,5.3,3],[3,3,5.3,5.3], alpha=0.09, color="#E53935")

ax2.axhline(3, color="#888", linewidth=1.2, linestyle="--", alpha=0.6)
ax2.axvline(3, color="#888", linewidth=1.2, linestyle="--", alpha=0.6)

for _, row in DMA.iterrows():
    size   = 260 if row["disclosed"] else 170
    alpha  = 0.90 if row["disclosed"] else 0.55
    marker = "o" if row["disclosed"] else "^"
    ax2.scatter(row["impact"], row["financial"],
                s=size, color=row["color"], alpha=alpha,
                marker=marker, edgecolors="white",
                linewidth=1.5, zorder=5)
    ax2.annotate(row["topic"],
                 xy=(row["impact"], row["financial"]),
                 xytext=(8, 4), textcoords="offset points",
                 fontsize=9, fontweight="bold", color=row["color"])

ax2.text(2.0, 5.18, "Impact material only",
         ha="center", fontsize=9, color="#FF7043",
         fontweight="bold", style="italic")
ax2.text(4.15, 1.12, "Financial material only",
         ha="center", fontsize=9, color="#1565C0",
         fontweight="bold", style="italic")
ax2.text(4.15, 5.18, "Both material",
         ha="center", fontsize=9, color="#C62828",
         fontweight="bold", style="italic")
ax2.text(1.85, 1.12, "Not material",
         ha="center", fontsize=9, color="#999999",
         fontweight="bold", style="italic")

ax2.set_xlabel("Impact Materiality\n(GE Vernova's impact on environment & society)",
               fontsize=11, color=DARK_TEXT, labelpad=10)
ax2.set_ylabel("Financial Materiality\n(sustainability risks & opportunities for GE Vernova)",
               fontsize=11, color=DARK_TEXT, labelpad=10)
ax2.set_xlim(1, 5.4); ax2.set_ylim(1, 5.4)
ax2.set_xticks([1,2,3,4,5])
ax2.set_yticks([1,2,3,4,5])
ax2.set_xticklabels(["1\nNot material","2\nLow","3\nThreshold",
                     "4\nHigh","5\nVery high"])
ax2.set_yticklabels(["1\nNot material","2\nLow","3\nThreshold",
                     "4\nHigh","5\nVery high"])
ax2.xaxis.grid(True, color=GRID_COL, linewidth=0.7)
ax2.yaxis.grid(True, color=GRID_COL, linewidth=0.7)
ax2.set_axisbelow(True)

p1 = mpatches.Patch(color="gray",       label="● Disclosed in 2025 report")
p2 = mpatches.Patch(color="gray", alpha=0.45, label="▲ Not yet disclosed (ESRS gap)")
ax2.legend(handles=[p1, p2], frameon=True, fontsize=9,
           loc="lower right", framealpha=0.9, edgecolor="#cccccc")

ax2.set_title(
    "Figure 2.  Illustrative Double Materiality Matrix — GE Vernova (ESRS 1 Framework)\n"
    "Which ESRS topics are material — and does GE Vernova currently disclose them?",
    fontsize=12, fontweight="bold", pad=14, color=DARK_TEXT)

fig2.text(0.01, 0.01,
    "Note: Illustrative DMA based on GE Vernova's business profile (energy technology, global operations). "
    "A formal DMA requires stakeholder engagement per ESRS 1, Art. 3. "
    "Threshold = 3.0. ● = disclosed in 2025 report. ▲ = gap.",
    fontsize=7.5, color=MID_TEXT, style="italic")

plt.tight_layout(rect=[0, 0.04, 1, 1])
plt.savefig(f"{OUTPUT_DIR}/fig2_double_materiality.png",
            dpi=150, bbox_inches="tight", facecolor="white")
plt.show(); plt.close()
print("Figure 2 saved.")


# ════════════════════════════════════════════════════════════════════════════
# FIGURE 3 — WHAT GE VERNOVA DISCLOSES VS WHAT ESRS REQUIRES
# ════════════════════════════════════════════════════════════════════════════

fig3, axes3 = plt.subplots(1, 2, figsize=(16, 7))
fig3.patch.set_facecolor("white")
fig3.suptitle(
    "Figure 3.  ESRS Topic Readiness — Scores & Priority Action Areas",
    fontsize=13, fontweight="bold", y=0.98, color=DARK_TEXT)

# Left: score by topic
ax3l = axes3[0]
ax3l.set_facecolor("#F8F8F8")
y = np.arange(len(ESRS))
bar_c = [GEV_GREEN if s>=7 else GEV_ORANGE if s>=4 else GEV_RED
         for s in ESRS["current_score"]]
bars = ax3l.barh(y, ESRS["current_score"], height=0.55,
                 color=bar_c, edgecolor="white", linewidth=0.8, alpha=0.85)
for bar, val in zip(bars, ESRS["current_score"]):
    ax3l.text(val+0.1, bar.get_y()+bar.get_height()/2,
              f"{val:.1f}", va="center", fontsize=9,
              fontweight="bold", color=DARK_TEXT)

ax3l.axvline(7, color=GEV_GREEN, linewidth=1.2,
             linestyle="--", alpha=0.6, label="Good (≥7)")
ax3l.axvline(4, color=GEV_ORANGE, linewidth=1.2,
             linestyle="--", alpha=0.6, label="Partial (≥4)")

ax3l.set_yticks(y)
ax3l.set_yticklabels(ESRS["topic"], fontsize=9, color=DARK_TEXT)
ax3l.set_xlabel("Current ESRS Score (0–10)", fontsize=10, color=DARK_TEXT)
ax3l.set_xlim(0, 12)
ax3l.xaxis.grid(True, color=GRID_COL, linewidth=0.7, zorder=0)
ax3l.set_axisbelow(True)
ax3l.legend(frameon=False, fontsize=9)
ax3l.set_title("(A)  Current ESRS Score per Topic",
               fontsize=11, fontweight="bold", pad=10, loc="left", color=DARK_TEXT)

# Right: radar / spider of gap
ax3r = axes3[1]
topics_short = ["E1","E2","E3","E4","E5","S1","S2","S3","S4","G1"]
N = len(topics_short)
angles = np.linspace(0, 2*np.pi, N, endpoint=False).tolist()
angles += angles[:1]

curr = ESRS["current_score"].tolist()
curr += curr[:1]
full = [10]*N + [10]

ax3r = fig3.add_subplot(122, polar=True)
ax3r.set_facecolor("white")

ax3r.plot(angles, full, color="#EEEEEE", linewidth=1)
ax3r.fill(angles, full, alpha=0.05, color="#EEEEEE")

ax3r.plot(angles, curr, color=GEV_BLUE, linewidth=2.2, marker="o", markersize=6)
ax3r.fill(angles, curr, alpha=0.18, color=GEV_BLUE)

ax3r.set_xticks(angles[:-1])
ax3r.set_xticklabels(topics_short, fontsize=10, fontweight="bold", color=DARK_TEXT)
ax3r.set_ylim(0, 10)
ax3r.set_yticks([2, 4, 6, 8, 10])
ax3r.set_yticklabels(["2","4","6","8","10"], fontsize=7.5, color=MID_TEXT)
ax3r.grid(color=GRID_COL, linewidth=0.7)
ax3r.set_title("(B)  ESRS Readiness Radar\n(outer ring = full compliance = 10)",
               fontsize=11, fontweight="bold", pad=20, color=DARK_TEXT)

fig3.text(0.01, 0.01,
    "Green = score ≥ 7 (good disclosure). Orange = 4–6 (partial). Red = < 4 (major gap). "
    "Radar shows all 10 ESRS topics. Source: GE Vernova Sustainability Report 2025.",
    fontsize=7.5, color=MID_TEXT, style="italic")

plt.savefig(f"{OUTPUT_DIR}/fig3_readiness.png",
            dpi=150, bbox_inches="tight", facecolor="white")
plt.show(); plt.close()
print("Figure 3 saved.")


# ════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ════════════════════════════════════════════════════════════════════════════

wb  = Workbook()
hdr_fill = PatternFill("solid", start_color="005EB8")
grn_fill = PatternFill("solid", start_color="E8F5E9")
red_fill = PatternFill("solid", start_color="FFEBEE")
ora_fill = PatternFill("solid", start_color="FFF3E0")
bold_wht = Font(name="Arial", bold=True, color="FFFFFF", size=10)
norm_blk = Font(name="Arial", color="1A1A1A", size=10)
center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
left     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin     = Side(style="thin", color="CCCCCC")
bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)

def make_hdr(ws, headers, widths):
    for c,(h,w) in enumerate(zip(headers,widths),1):
        cell=ws.cell(1,c,h)
        cell.font=bold_wht; cell.fill=hdr_fill
        cell.alignment=center; cell.border=bdr
        ws.column_dimensions[get_column_letter(c)].width=w
    ws.row_dimensions[1].height=30

# Sheet 1: ESRS Gap Analysis
ws1=wb.active; ws1.title="ESRS Gap Analysis"
ws1.sheet_view.showGridLines=False
make_hdr(ws1,
    ["ESRS Topic","Pillar","Current Score\n(0–10)","Gap\n(0–10)",
     "Priority","What GE Vernova Discloses","Primary CSRD Gap"],
    [26,10,16,12,12,50,52])
for r,row in ESRS.iterrows():
    sc=row["current_score"]
    fill=grn_fill if sc>=7 else (ora_fill if sc>=4 else red_fill)
    for c,v in enumerate([row["topic"],row["pillar"],sc,row["gap_score"],
        row["priority"],row["what_gev_discloses"],row["key_gap"]],1):
        cell=ws1.cell(r+2,c,v)
        cell.font=norm_blk; cell.fill=fill
        cell.alignment=center if c not in [1,6,7] else left
        cell.border=bdr
    ws1.row_dimensions[r+2].height=50

# Sheet 2: Double Materiality
ws2=wb.create_sheet("Double Materiality")
ws2.sheet_view.showGridLines=False
make_hdr(ws2,
    ["ESRS Topic","Impact Score (1–5)","Financial Score (1–5)",
     "Both Material?","Currently Disclosed?"],
    [26,22,22,18,20])
for r,row in DMA.iterrows():
    both = (row["impact"]>=3) and (row["financial"]>=3)
    fill=grn_fill if row["disclosed"] else (red_fill if both else ora_fill)
    for c,v in enumerate([row["topic"],row["impact"],row["financial"],
        "✓ Yes" if both else "—",
        "✓ Yes" if row["disclosed"] else "✗ No"],1):
        cell=ws2.cell(r+2,c,v)
        cell.font=norm_blk; cell.fill=fill
        cell.alignment=center if c!=1 else left
        cell.border=bdr

# Sheet 3: Methodology
ws3=wb.create_sheet("Methodology")
ws3.sheet_view.showGridLines=False
ws3.column_dimensions["A"].width=32
ws3.column_dimensions["B"].width=65
meta=[
    ("Field","Detail"),
    ("Regulation","CSRD — Corporate Sustainability Reporting Directive (EU) 2022/2464"),
    ("Standards","ESRS Set 1 (2023) — European Sustainability Reporting Standards"),
    ("Key Article","ESRS 1, Art. 3 — Double Materiality Assessment (DMA)"),
    ("DMA Definition","Impact materiality + financial materiality"),
    ("Threshold","Score ≥ 3.0 on either dimension = material topic"),
    ("Scoring","0 = not disclosed, 5 = partial, 10 = full ESRS compliance"),
    ("Company","GE Vernova — energy technology (wind, gas, grid, nuclear)"),
    ("Data source","GE Vernova Sustainability Report 2025"),
    ("Frameworks aligned","GRI, SASB, TCFD, UN SDGs (current); ESRS (gap identified)"),
    ("Analysis by","Matin Bahadori — M.Sc. Umwelttechnik, BTU Cottbus-Senftenberg"),
]
for r,(a,b) in enumerate(meta,1):
    ca=ws3.cell(r,1,a); cb=ws3.cell(r,2,b)
    is_hdr=r==1
    ca.font=Font(name="Arial",bold=True,color="FFFFFF" if is_hdr else "1A1A1A",size=10)
    cb.font=Font(name="Arial",color="FFFFFF" if is_hdr else "1A1A1A",size=10,bold=is_hdr)
    if is_hdr: ca.fill=hdr_fill; cb.fill=hdr_fill
    ca.alignment=left; cb.alignment=left
    ca.border=bdr; cb.border=bdr

path_xl=f"{OUTPUT_DIR}/gev_csrd_gap.xlsx"
wb.save(path_xl)
print(f"Excel saved → {path_xl}")

print(f"\n{'='*60}")
print(f"  GE Vernova — CSRD/ESRS Gap Summary")
print(f"{'='*60}")
print(f"  Overall ESRS readiness  : {ESRS['current_score'].mean():.1f}/10")
print(f"  Strongest topic         : {ESRS.loc[ESRS.current_score.idxmax(),'topic']} ({ESRS.current_score.max():.1f}/10)")
print(f"  Weakest topic           : {ESRS.loc[ESRS.current_score.idxmin(),'topic']} ({ESRS.current_score.min():.1f}/10)")
print(f"  Topics with gap > 5     : {(ESRS.gap_score>5).sum()}")
print(f"  #1 missing requirement  : Double Materiality Assessment (ESRS 1 Art. 3)")
print(f"  Both material topics    : {((DMA.impact>=3)&(DMA.financial>=3)).sum()}/{len(DMA)}")
print(f"{'='*60}")
